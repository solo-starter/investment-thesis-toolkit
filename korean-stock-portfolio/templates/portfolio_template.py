"""
한국 주식 포트폴리오 엑셀 생성 템플릿
=========================================

사용 방법:
1. 데이터 딕셔너리 (PORTFOLIO_DATA, EARNINGS_DATA, VALUATION_DATA)를 채운다
2. 기존 파일이 있으면 load_workbook으로 열고, 없으면 Workbook() 새로 생성
3. 6개 시트를 갱신 (히스토리는 추가)
4. /mnt/user-data/outputs/ 에 저장

이 템플릿은 분기별로 같은 코드 구조를 재사용하기 위한 참조용이다.
실제 사용 시 데이터를 새로 수집해서 채워 넣는다.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime

# ============================================================
# 1. 표준 스타일 정의 (분기마다 동일)
# ============================================================

STYLES = {
    'header_font': Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF"),
    'header_fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
    'title_font': Font(name="맑은 고딕", size=14, bold=True, color="1F4E78"),
    'subtitle_font': Font(name="맑은 고딕", size=9, italic=True, color="666666"),
    'normal_font': Font(name="맑은 고딕", size=10),
    'bold_font': Font(name="맑은 고딕", size=10, bold=True),

    # 섹터별 색상
    'sector_semi': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),
    'sector_ai': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    'sector_robot': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),

    # 시나리오별 색상
    'cons_fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    'base_fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    'opt_fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    'rerate_fill': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),

    # 상승/하락
    'upside_pos': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'upside_neg': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SECTOR_COLOR_MAP = {
    "반도체": STYLES['sector_semi'],
    "AI 핵심전력 인프라": STYLES['sector_ai'],
    "피지컬 AI 로봇": STYLES['sector_robot'],
}

# ============================================================
# 2. 적정가 산정 함수
# ============================================================

def calculate_fair_prices(item):
    """
    종목 데이터에서 4단계 시나리오별 적정가 계산
    
    item: dict
        - method: "PER × EPS" | "NAV 할인" | "PSR × Sales"
        - fwd_eps: int (PER 방식) | None (NAV/PSR)
        - per_cons, per_base, per_opt, per_rerate: 멀티플
        - nav_per_share: int (NAV 방식만)
        - discount_rates: list (NAV 방식만, 보수/기준/낙관/재평가 순)
        - sales_per_share: int (PSR 방식만)
        - psr_multiples: list (PSR 방식만)
    
    반환: (fair_prices_list, rerate_price)
    """
    method = item["method"]
    
    if method == "PER × EPS" and isinstance(item.get("fwd_eps"), (int, float)):
        eps = item["fwd_eps"]
        fairs = [int(eps * item[k]) for k in ["per_cons", "per_base", "per_opt"]]
        rerate = int(eps * item["per_rerate"])
        return fairs, rerate
    
    elif method == "NAV 할인":
        nav = item["nav_per_share"]
        rates = item["discount_rates"]  # [보수, 기준, 낙관, 재평가]
        fairs = [int(nav * (1 - r)) for r in rates[:3]]
        rerate = int(nav * (1 - rates[3]))
        return fairs, rerate
    
    elif method == "PSR × Sales":
        sps = item["sales_per_share"]
        psrs = item["psr_multiples"]  # [보수, 기준, 낙관, 재평가]
        fairs = [int(sps * p) for p in psrs[:3]]
        rerate = int(sps * psrs[3])
        return fairs, rerate
    
    return ["—", "—", "—"], "—"


# ============================================================
# 3. 시트 생성 함수들
# ============================================================

def create_portfolio_sheet(wb, data, ref_date):
    """포트폴리오 종합 시트"""
    if "포트폴리오 종합" in wb.sheetnames:
        del wb["포트폴리오 종합"]
    ws = wb.create_sheet("포트폴리오 종합", 0)
    
    ws["A1"] = "한국 주식 포트폴리오 관리"
    ws["A1"].font = STYLES['title_font']
    ws.merge_cells("A1:L1")
    ws["A1"].alignment = CENTER
    
    ws["A2"] = f"기준일: {ref_date} | 통화: KRW"
    ws["A2"].font = STYLES['subtitle_font']
    ws.merge_cells("A2:L2")
    ws["A2"].alignment = CENTER
    
    headers = ["섹터", "종목명", "종목코드", "거래소", "현재가", "시가총액",
               "52주 최고", "52주 최저", "PER(TTM)", "EPS(TTM)", "배당수익률", "투자 메모"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = STYLES['header_font']
        c.fill = STYLES['header_fill']
        c.alignment = CENTER
        c.border = BORDER
    
    for ri, row in enumerate(data, start=5):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = STYLES['normal_font']
            cell.border = BORDER
            # ... (열별 스타일 적용)
    
    col_widths = {"A": 18, "B": 18, "C": 11, "D": 8, "E": 13, "F": 11,
                  "G": 13, "H": 12, "I": 10, "J": 11, "K": 11, "L": 60}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C5"
    return ws


def append_history_row(wb, quarter_label, data):
    """
    분기별 히스토리 시트에 새 행 추가 (덮어쓰기 X, 누적)
    
    quarter_label: "2026 1Q", "2026 2Q" 형식
    data: list of dict (종목별)
    """
    sheet_name = "분기별 히스토리"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # 헤더 작성
        ws["A1"] = "📈 분기별 적정가 히스토리"
        ws["A1"].font = STYLES['title_font']
        ws.merge_cells("A1:H1")
        ws["A1"].alignment = CENTER
        
        ws["A2"] = "매분기 종목별 핵심 지표를 한 줄씩 누적. 추세 분석용."
        ws["A2"].font = STYLES['subtitle_font']
        ws.merge_cells("A2:H2")
        
        headers = ["분기", "섹터", "종목", "현재가", "기준 적정가", 
                   "상승여력(%)", "컨센서스 목표가", "메모"]
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = STYLES['header_font']
            c.fill = STYLES['header_fill']
            c.alignment = CENTER
            c.border = BORDER
        ws.freeze_panes = "A5"
    else:
        ws = wb[sheet_name]
    
    # 다음 빈 행 찾기
    next_row = ws.max_row + 1
    
    # 데이터 행 추가
    for item in data:
        upside = (item['fair_base'] - item['price']) / item['price'] if item['fair_base'] else None
        ws.cell(row=next_row, column=1, value=quarter_label).alignment = CENTER
        ws.cell(row=next_row, column=2, value=item['sector']).fill = SECTOR_COLOR_MAP[item['sector']]
        ws.cell(row=next_row, column=3, value=f"{item['name']} ({item['code']})")
        ws.cell(row=next_row, column=4, value=item['price']).number_format = "#,##0"
        ws.cell(row=next_row, column=5, value=item['fair_base']).number_format = "#,##0"
        if upside is not None:
            uc = ws.cell(row=next_row, column=6, value=upside)
            uc.number_format = "+0.0%;-0.0%;0.0%"
            if upside > 0.10:
                uc.fill = STYLES['upside_pos']
            elif upside < -0.10:
                uc.fill = STYLES['upside_neg']
        ws.cell(row=next_row, column=7, value=item.get('consensus_tp', '—')).number_format = "#,##0"
        ws.cell(row=next_row, column=8, value=item.get('memo', ''))
        
        for ci in range(1, 9):
            ws.cell(row=next_row, column=ci).border = BORDER
            if ci != 8:
                ws.cell(row=next_row, column=ci).font = STYLES['normal_font']
        
        next_row += 1
    
    # 컬럼 너비
    col_widths = {"A": 10, "B": 18, "C": 22, "D": 12, "E": 13, "F": 13, "G": 14, "H": 40}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w


# ============================================================
# 4. 메인 실행 흐름 (의사 코드)
# ============================================================

def update_portfolio(existing_file=None, new_data=None, quarter_label=None):
    """
    포트폴리오 갱신 메인 함수
    
    existing_file: 기존 엑셀 파일 경로 (없으면 신규 생성)
    new_data: dict {
        'portfolio': [...],  # 포트폴리오 종합 데이터
        'earnings': [...],    # 실적 발표 일정
        'financials': [...],  # 재무지표
        'sector_memo': [...], # 섹터 분석
        'valuation': [...],   # 적정가 분석
        'ref_date': 'YYYY-MM-DD'
    }
    quarter_label: "2026 1Q" 형식 (히스토리 추가용)
    """
    # 1. 파일 로드 또는 신규 생성
    if existing_file and os.path.exists(existing_file):
        wb = load_workbook(existing_file)
    else:
        wb = Workbook()
        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    
    # 2. 메인 5개 시트 갱신 (덮어쓰기)
    create_portfolio_sheet(wb, new_data['portfolio'], new_data['ref_date'])
    # create_earnings_sheet(wb, new_data['earnings'])
    # create_financials_sheet(wb, new_data['financials'])
    # create_sector_sheet(wb, new_data['sector_memo'])
    # create_valuation_sheet(wb, new_data['valuation'])
    
    # 3. 히스토리 시트는 추가 (누적)
    if quarter_label:
        append_history_row(wb, quarter_label, new_data['valuation'])
    
    # 4. 저장
    out_path = "/mnt/user-data/outputs/한국주식_포트폴리오.xlsx"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


# ============================================================
# 5. 데이터 구조 예시
# ============================================================

EXAMPLE_VALUATION_ITEM = {
    "sector": "반도체",
    "name": "삼성전자",
    "code": "005930",
    "price": 268500,
    "method": "PER × EPS",
    "fwd_eps": 35000,
    "per_cons": 8,
    "per_base": 11,
    "per_opt": 14,
    "per_rerate": 18,
    "consensus_tp": 274603,
    "note": "..."
}

EXAMPLE_NAV_ITEM = {
    "sector": "반도체",
    "name": "SK스퀘어",
    "code": "402340",
    "price": 1090000,
    "method": "NAV 할인",
    "nav_per_share": 2000000,
    "discount_rates": [0.40, 0.30, 0.20, 0.10],  # 보수/기준/낙관/재평가
    "consensus_tp": 770000,
    "note": "..."
}

EXAMPLE_PSR_ITEM = {
    "sector": "피지컬 AI 로봇",
    "name": "두산로보틱스",
    "code": "454910",
    "price": 91100,
    "method": "PSR × Sales",
    "sales_per_share": 1300,
    "psr_multiples": [8, 15, 25, 40],  # 보수/기준/낙관/재평가
    "consensus_tp": 109000,
    "note": "..."
}

EXAMPLE_HISTORY_ITEM = {
    "sector": "반도체",
    "name": "삼성전자",
    "code": "005930",
    "price": 268500,
    "fair_base": 385000,  # 기준 시나리오 적정가
    "consensus_tp": 274603,
    "memo": "1Q26 영업이익 컨센서스 상회"
}

# ============================================================
# 메모: 이 템플릿은 참조용 코드 구조다.
# 실제 분기별 갱신 시:
# 1. 새 종목 데이터를 웹 검색으로 수집
# 2. 위 함수 구조에 맞게 데이터 채우기
# 3. update_portfolio() 호출
# ============================================================
